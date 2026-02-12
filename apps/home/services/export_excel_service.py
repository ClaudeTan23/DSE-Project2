from django.http import HttpResponse
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def export(data):

    df = pd.DataFrame(data)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="report_{timestamp}.xlsx"'
    )

    # Write Excel using openpyxl engine
    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")

        ws = writer.sheets["Report"]

        header_fill = PatternFill("solid", fgColor="BDD7EE") 
        header_font = Font(bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        ws.auto_filter.ref = ws.dimensions

        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

    return response
